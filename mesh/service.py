import logging

from asynch.cursors import DictCursor
from openpyxl.styles import Font, Alignment
from openpyxl.workbook import Workbook

from mesh.model import GeographyModel
from settings import clickhouse_holder
from utils.utils import ResponseObject, date_condition_from_segment


class MeshService:
    class CategoryConfig:
        def __init__(self, r_name, table):
            self.r_name = r_name
            self.table = table

    __schema = 'mesh'
    categories_config_map = {
        'visits': CategoryConfig('Визиты', 'sessions'),
        'uniq_users': CategoryConfig('Уникальные пользователи', 'uniqusers'),
        'devices': CategoryConfig('Тип устройства', 'devices'),
        'os': CategoryConfig('Операционная система устройства', 'os'),
        'ages': CategoryConfig('Возраст', 'ages'),
        'geography': CategoryConfig('География', 'ao')
    }
    filters_map = {
        'segment': 'Сегмент',
        'date_from': 'Начало периода',
        'date_to': 'Конец периода',
        'product': 'Продукт',
        'g_type': 'Тип'
    }

    @classmethod
    async def get_date_ranges(cls, category):
        if category not in cls.categories_config_map:
            return ResponseObject(meta=f'Для категории {category} не задана таблица')
        sql = f'select  min(from_dt) as min,  max(from_dt) as max from {cls.__schema}.{cls.categories_config_map[category].table}'
        async with await clickhouse_holder.get_connection() as conn:
            async with conn.cursor(cursor=DictCursor) as cursor:
                await cursor.execute(sql)
                return ResponseObject(data=await cursor.fetchone())

    @classmethod
    async def get_aos_and_districts(cls):
        source = f'{cls.__schema}.{cls.categories_config_map['geography'].table}'
        sql = (f'select distinct AO, District from {source} where '
               f'from_dt  = (select max(from_dt) from {source}) order by AO')
        async with await clickhouse_holder.get_connection() as conn:
            async with conn.cursor(cursor=DictCursor) as cursor:
                await cursor.execute(sql)
                data = await cursor.fetchall()
                aos_districts = {k['AO']: list() for k in data}
                for row in data:
                    aos_districts[row['AO']].append(row['District'])
                return ResponseObject(data=[{
                    'ao_name': k,
                    'districts': v
                } for k, v in aos_districts.items()])

    @classmethod
    async def __get_data_type_1(cls, sql, meta_value_name, column_value_name):
        async with await clickhouse_holder.get_connection() as conn:
            async with conn.cursor(cursor=DictCursor) as cursor:
                await cursor.execute(sql)
                data = await cursor.fetchall()
                if len(data) == 0:
                    return ResponseObject(meta='Не получены строки с данными')
                response = ResponseObject(
                    data=list(),
                    meta={
                        'section_values': set(),
                        meta_value_name: set()
                    }
                )
                response.meta['date'] = data[0]['from_dt']
                for row in data:
                    response.meta['section_values'].add(row['section'])
                    response.meta[meta_value_name].add(row[column_value_name])
                    response.data.append({k: v for k, v in row.items() if k != 'from_dt'})
                response.meta[meta_value_name] = sorted(response.meta[meta_value_name])
                return response

    @classmethod
    async def get_ages_data(cls):
        source = f'{cls.__schema}.{cls.categories_config_map['ages'].table}'
        sql = f'select * from {source} where from_dt  = (select max(from_dt) from {source}) order by section, age_range'
        return await cls.__get_data_type_1(sql, 'age_range_values', 'age_range')

    @classmethod
    async def get_os_data(cls):
        source = f'{cls.__schema}.{cls.categories_config_map['os'].table}'
        sql = f'select * from {source} where from_dt  = (select max(from_dt) from {source}) order by section, user_agent_os_family'
        return await cls.__get_data_type_1(sql, 'user_agent_os_family_values', 'user_agent_os_family')

    @classmethod
    async def get_devices_data(cls):
        source = f'{cls.__schema}.{cls.categories_config_map['devices'].table}'
        sql = f'select * from {source} where from_dt  = (select max(from_dt) from {source}) order by section, device'
        return await cls.__get_data_type_1(sql, 'device_values', 'device')

    @classmethod
    async def __get_data_type_2(cls, source, product, segment, date_from, date_to):
        sql = (f'select from_dt, kol_vo from {source} where section = %(product)s and agg_type = %(segment)s and'
               f' {date_condition_from_segment(segment)} order by from_dt')
        params = {
            'product': product,
            'segment': segment,
            'date_from': date_from,
            'date_to': date_to
        }
        async with await clickhouse_holder.get_connection() as conn:
            async with conn.cursor(cursor=DictCursor) as cursor:
                await cursor.execute(sql, params)
                data = await cursor.fetchall()
                return ResponseObject(data=[{
                    'date': d['from_dt'],
                    'value': d['kol_vo']
                } for d in data])

    @classmethod
    async def get_visits_data(cls, product, segment, date_from, date_to):
        source = f'{cls.__schema}.{cls.categories_config_map['visits'].table}'
        return await cls.__get_data_type_2(source, product, segment, date_from, date_to)

    @classmethod
    async def get_uniq_users_data(cls, product, segment, date_from, date_to):
        source = f'{cls.__schema}.{cls.categories_config_map['uniq_users'].table}'
        return await cls.__get_data_type_2(source, product, segment, date_from, date_to)

    @classmethod
    async def get_geography_data(cls, g_type, g_name):
        source = f'{cls.__schema}.{cls.categories_config_map['geography'].table}'
        sql = (f'select section, :kol_vo, :type_fields from {source} where from_dt  = (select max(from_dt) '
               f'from {source}) :current_type :group_by order by :type_fields')
        sql = sql.replace(':current_type', 'and :current_type = %(current_value)s' if g_name else '')
        sql = sql.replace(':kol_vo',
                          'kol_vo' if g_type == 'district' else 'sum(kol_vo) kol_vo' if g_type == 'ao' else '')
        sql = sql.replace(':current_type', 'District' if g_type == 'district' else 'AO' if g_type == 'ao' else '')
        sql = sql.replace(':type_fields', 'District, AO' if g_type == 'district' else 'AO' if g_type == 'ao' else '')
        sql = sql.replace(':group_by', '' if g_type == 'district' else 'group by section, AO' if g_type == 'ao' else '')
        meta = {'data_values': set()}
        async with await clickhouse_holder.get_connection() as conn:
            async with conn.cursor(cursor=DictCursor) as cursor:
                await cursor.execute(sql, {'current_value': g_name} if g_name else None)
                rows = await cursor.fetchall()
                if len(rows) == 0:
                    return ResponseObject(meta='Не получены строки с данными')
                if g_name:
                    result = GeographyModel(ao_name=rows[0]['AO'])
                    if g_type == 'district':
                        result.district_name = g_name
                        result.district_data = dict()
                    elif g_type == 'ao':
                        result.ao_data = dict()
                    for row in rows:
                        meta['data_values'].add(row['section'])
                        if g_type == 'district':
                            result.district_data[row['section']] = row['kol_vo']
                        elif g_type == 'ao':
                            result.ao_data[row['section']] = row['kol_vo']
                    return ResponseObject(data=result.model_dump(exclude_none=True), meta=meta)
                else:
                    meta[f'{g_type}_name_values'] = set()
                    names_column = 'District' if g_type == 'district' else 'AO' if g_type == 'ao' else ''
                    all_names = {row[names_column] for row in rows}
                    data = {n: GeographyModel(district_data=dict()) if g_type == 'district' else GeographyModel(
                        ao_data=dict())
                            for n in all_names}
                    for row in rows:
                        meta[f'{g_type}_name_values'].add(row[names_column])
                        meta['data_values'].add(row['section'])
                        data[row[names_column]].ao_name = row['AO']
                        if g_type == 'district':
                            data[row[names_column]].district_name = row[names_column]
                            data[row[names_column]].district_data[row['section']] = row['kol_vo']
                        elif g_type == 'ao':
                            data[row[names_column]].ao_data[row['section']] = row['kol_vo']
                    meta[f'{g_type}_name_values'] = sorted(meta[f'{g_type}_name_values'])
                    return ResponseObject(data=[data[k].model_dump(exclude_none=True) for k in sorted(data)], meta=meta)

    @classmethod
    async def report_download(cls, filter):
        wb = Workbook()
        ws = wb.active
        category = filter.pop('category')
        ws.title = cls.categories_config_map.get(category).r_name
        header_font = Font(size=12, bold=True)
        value_font = Font(size=12, bold=False)
        wrap_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        row_num = 1

        def __set_cell(cell, value, font, alignment):
            cell.value = value
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment

        if category == 'visits' or category == 'uniq_users':
            for f_name in filter:
                filter_name_cell = ws.cell(row=row_num, column=1)
                __set_cell(filter_name_cell, cls.filters_map[f_name] + ': ', header_font, wrap_alignment)

                filter_value_cell = ws.cell(row=row_num, column=2)
                cell_value = None
                if f_name == 'date_from' or f_name == 'date_to':
                    cell_value = filter[f_name].strftime('%d.%m.%Y')
                elif f_name == 'segment':
                    match filter[f_name]:
                        case 'date':
                            cell_value = 'День'
                        case 'year':
                            cell_value = 'Год'
                        case 'month':
                            cell_value = 'Месяц'
                        case 'weekofyear':
                            cell_value = 'Неделя'
                else:
                    cell_value = filter[f_name]
                __set_cell(filter_value_cell, cell_value, value_font, wrap_alignment)
                row_num += 1
        elif category in ['devices', 'os', 'ages', 'geography']:
            filter_name_cell = ws.cell(row=row_num, column=1)
            __set_cell(filter_name_cell, 'Дата: ', header_font, wrap_alignment)

            filter_value_cell = ws.cell(row=row_num, column=2)
            __set_cell(filter_value_cell, (await cls.get_date_ranges(category)).data['max'].strftime('%d.%m.%Y'),
                           value_font, wrap_alignment)

            if category == 'geography':
                row_num += 1
                filter_name_cell = ws.cell(row=row_num, column=1)
                __set_cell(filter_name_cell, 'Тип: ', header_font, wrap_alignment)

                filter_value_cell = ws.cell(row=row_num, column=2)
                __set_cell(filter_value_cell,
                               'Районы' if filter['g_type'] == 'district' else 'АО' if filter['g_type'] == 'ao'
                               else filter['type'], value_font, wrap_alignment)
        row_num += 2

        def __create_table_type1(table_data, row_num=row_num):
            data_header_cell = ws.cell(row=row_num, column=1)
            count_header_cell = ws.cell(row=row_num, column=2)
            __set_cell(data_header_cell, 'Дата: ', header_font, wrap_alignment)
            __set_cell(count_header_cell, 'Количество', header_font, wrap_alignment)
            ws.column_dimensions[data_header_cell.column_letter].width = 20
            ws.column_dimensions[count_header_cell.column_letter].width = 20
            row_num += 1
            for d in table_data:
                data_value_cell = ws.cell(row=row_num, column=1)
                count_value_cell = ws.cell(row=row_num, column=2)
                __set_cell(data_value_cell, d['date'].strftime('%d.%m.%Y'), value_font, wrap_alignment)
                __set_cell(count_value_cell, d['value'], value_font, wrap_alignment)
                row_num += 1

        match category:
            case 'uniq_users':
                __create_table_type1((await cls.get_uniq_users_data(filter['product'], filter['segment'],
                                                                    filter['date_from'], filter['date_to'])).data)
            case 'visits':
                __create_table_type1((await cls.get_visits_data(filter['product'], filter['segment'],
                                                                    filter['date_from'], filter['date_to'])).data)
        return wb
